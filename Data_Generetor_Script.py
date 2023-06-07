Import necessary libraries
import pandas as pd
import random
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import string
Generate data for "Clients" sheet.

clients = []
for i in range(1, 1501):
    clientes = {
        "client_id": ''.join(random.choices(string.digits, k=7)),
        "name": random.choice(['Juan', 'María', 'Pedro', 'Laura', 'Gabriela', 'José', 'Carla', 'Antonio', 'Ana', 'Jorge', 'Valeria', 'Miguel', 'Cristina', 'Fernando', 'Julia', 'Ricardo', 'Renata', 'Diego', 'Sofía', 'Daniel']),
        "last_name": random.choice(['González', 'Pérez', 'Martínez', 'Sánchez', 'López', 'Gómez', 'Hernández', 'Fernández', 'Rodríguez', 'García', 'Díaz', 'Torres', 'Ramos', 'Ruiz', 'Moreno', 'Alonso', 'Romero', 'Jiménez', 'Álvarez', 'Vargas']),
        "addresses": f"{random.choice(['Calle','Carrera'])} {random.randint(1, 100)} #{random.randint(1, 20)}-{random.randint(1, 100)}",
        "city": random.choice(["Bogotá", "Medellín", "Cali", "Cartagena", "Ibagué"]),
        "Country": "Colombia"
    }
    clients.append(clientes)
Create a DataFrame to clients
df_clients = pd.DataFrame(clients)

#Top 3 rows of dataframe
df_clients.head(3)

#Bottom 3 rows of dataframe
#df_clients.tail(3)
Data type changes
#Changes
df_clients['client_id'] = df_clients['client_id'].astype('Int64')

#DataFrame overview
df_clients.info()
Generate data for "Products" sheet.

products = []
for i in range(1, 21):
    unit_price = random.uniform(10, 100)
    discount_rate = random.uniform(1, 50)
    producto = {
        "product_id": ''.join(random.choices(string.ascii_uppercase, k=3)) + '-' + ''.join(random.choices(string.ascii_lowercase + string.digits, k=9)),
        "product_name": f"Producto {i}",
        "Description": f"Product Description {i}.",
        "category": random.choice(["Electrónica", "Ropa", "Hogar", "Deportes", "Juguetes"]),
        "product_weight": ("%.2f" % random.uniform(0, 20)),
        "unit_price": ("%.2f" % unit_price),
        "Unit_Cost": ("%.2f" % (unit_price * (1 - (discount_rate/100))))
    }
    products.append(producto)

    
Create a DataFrame to products
df_products = pd.DataFrame(products)

#Top 3 rows of dataframe
df_products.head(3)
Data type changes
#Changes
df_products['product_weight'] = df_products['product_weight'].astype(float)
df_products['unit_price'] = df_products['unit_price'].astype(float)
df_products['Unit_Cost'] = df_products['Unit_Cost'].astype(float)

#DataFrame overview
df_products.info()
Generate data for "Ordes" sheet.
orders = []
for i in range(1, 35001):
    start_date = datetime.date(2018, 1, 1)
    end_date = datetime.date.today()
    days_between = (end_date - start_date).days
    delta = end_date - start_date
    date = start_date + datetime.timedelta(days=random.randint(0, delta.days))
    status = random.choice(["pendiente","pagado","rechazado"])
    delivery_status = random.choice(["pendiente","enviado","entregado"]) if status == "pagado" else "pendiente" if status == "pendiente" else "rechazado"
    pedido = {
        "Order_id": i,
        "Order_date": date,
        "status": status,
        "delivery_status": delivery_status,
        "client_id": random.choice(clients)["client_id"],
        "product_id": random.choice(products)["product_id"],
        "items": random.randint(1, 5),
        "discount_amount": ("%.2f" % random.uniform(0, 10))
    }
    orders.append(pedido)
Create a DataFrame to orders
df_orders = pd.DataFrame(orders)

#Top 3 rows of dataframe
df_orders.head(3)
Data type changes
#Changes
df_orders['Order_date'] = pd.to_datetime(df_orders['Order_date'])
df_orders['client_id'] = df_orders['client_id'].astype('Int64')
df_orders['discount_amount'] = df_orders['discount_amount'].astype(float)

#DataFrame overview
df_orders.info()
df_orders.describe()

DataFrame's dimensions
[df_clients.ndim,df_products.ndim,df_orders.ndim]
Create a excel workbook
workbook = Workbook()
Insert each DataFrame in a separate Excel sheets
workbook.create_sheet("Clients")
sheet = workbook["Clients"]
for row in dataframe_to_rows(df_clients, index=False, header=True):
    sheet.append(row)

workbook.create_sheet("Products")
sheet = workbook["Products"]
for row in dataframe_to_rows(df_products, index=False, header=True):
    sheet.append(row)

workbook.create_sheet("Orders")
sheet = workbook["Orders"]
for row in dataframe_to_rows(df_orders, index=False, header=True):
    sheet.append(row)
Save and close workbook
workbook.save(filename="Raw_Data.xlsx")
